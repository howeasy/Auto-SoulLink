#!/usr/bin/env python3
"""
gen_rr_priority_trainers.py — Parse the community-maintained
"Pokémon Radical Red 4.1 — Default Mode" trainer spreadsheet into
data/games/gen3_frlge/rr_priority_trainers.json.

Source workbook (publicly viewable):
    https://docs.google.com/spreadsheets/d/1ES8L4OzeJ8rCuMWFNvrDaZKArqR7Vys2ytFxjx2pbwE

Usage:
    # 1) Download the xlsx export of the sheet to a local path (default
    #    location is repo root):
    curl -sL "https://docs.google.com/spreadsheets/d/1ES8L4OzeJ8rCuMWFNvrDaZKArqR7Vys2ytFxjx2pbwE/export?format=xlsx" \
        -o rr_trainers_dump.xlsx
    # 2) Run the parser. Output lands in data/games/gen3_frlge/.
    python tools/gen_rr_priority_trainers.py [--src rr_trainers_dump.xlsx]

Each trainer entry on the "boss" sheets is a vertically-stacked block:
  +0  : Pokémon nickname/species (Geodude-A, Kleavor, etc.)
  +1  : Level (e.g. "Max Level - 2" or numeric)
  +2  : (gender / type / blank)
  +3  : (Tera type / blank)
  +4  : Nature
  +5  : Ability
  +6  : Held item
  +7…+10 : 4 moves
Mon blocks repeat across columns at cols 5, 10, 15, 20, 25, 30 (5-col stride);
col 3 in the same header row holds the trainer class+name as
"CLASS\\nNAME" (e.g. "GYM LEADER\\nBROCK", "RIVAL\\nTERRY", "MINI BOSS\\nKOMODO").

Trainer Order sheet provides the recommended in-game progression
(trainer name → area name pairs, with "(OPTIONAL)" markers between).
"""

import argparse
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_SCRIPT_DIR = Path(__file__).resolve().parent
_REPO_ROOT  = _SCRIPT_DIR.parent
_DATA_DIR   = _REPO_ROOT / "data" / "games" / "gen3_frlge"
_OUT_PATH   = _DATA_DIR / "rr_priority_trainers.json"
_RR_TRAINERS_PATH = _DATA_DIR / "rr_trainers.json"
_CALC_NORMAL_PATH    = _REPO_ROOT / "calc" / "src" / "js" / "data" / "sets" / "normal.js"
_CALC_SUPP_SRC_PATH  = _REPO_ROOT / "calc" / "src"  / "js" / "data" / "sets" / "slink_priority.js"
_CALC_SUPP_DIST_PATH = _REPO_ROOT / "calc" / "dist" / "js" / "data" / "sets" / "slink_priority.js"

# Species name normalisation: sheet uses some abbreviated forms (e.g.
# "Geodude-A", "Growlithe-H") that need expanding so they match the calc
# pokédex keys (window.pokedex / window.SETDEX_SV).
_SHEET_SPECIES_MAP = {
    # Regional suffixes
    "-A":  "-Alola",
    "-H":  "-Hisui",
    "-G":  "-Galar",
    "-P":  "-Paldea",
}


def _normalize_species(s: str) -> str:
    """Expand short suffixes ("Geodude-A" → "Geodude-Alola") and tidy spacing."""
    s = (s or "").strip()
    if not s:
        return s
    for short, long in _SHEET_SPECIES_MAP.items():
        if s.endswith(short):
            return s[: -len(short)] + long
    return s


def _synthesise_calc_label(cls: str, name: str, fight_label: str) -> str:
    """Build a stable canonical set name for the calc Prep tab.

    Examples:
      ("Gym Leader", "Brock", "")               → "Gym Leader Brock"
      ("Leader",     "Bugsy", "Pre Lt. Surge")  → "Leader Bugsy (Pre Lt. Surge)"
      ("Rival",      "Rival", "If Rival Has Squirtle")
                                                → "Rival (If Rival Has Squirtle)"
    """
    cls = (cls or "").strip()
    name = (name or "").strip()
    fl = (fight_label or "").strip()
    base = f"{cls} {name}".strip()
    # Avoid "Rival Rival" — when class == name, collapse to one.
    if cls.lower() == name.lower():
        base = name
    if fl:
        return f"{base} ({fl})"
    return base

# Sheets that follow the standard boss-block layout (5-col-stride trainer blocks)
_BOSS_SHEETS = [
    "Kanto Leaders",
    "Kanto Rematch",
    "Johto Leaders",
    "Indigo League",
    "Rivals",
    "Team Rocket",
    "Mini Bosses",
    "Optional Bosses",
    "Postgame",
]

# Column strides where trainer mon data lives.
_MON_COLS = [5, 10, 15, 20, 25, 30]

# Offsets relative to the trainer header row.
_OFF_LEVEL   = 1
_OFF_NATURE  = 4
_OFF_ABILITY = 5
_OFF_ITEM    = 6
_OFF_MOVE_0  = 7

# Class/name header column (1-based in openpyxl indexing).
_HEADER_COL = 3


def _norm_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


# Spreadsheet calls the player's rival "Rival" — RR's actual trainer name
# (in rr_trainers.json) is "Terry". Same idea would apply to any other
# generic-named slots if they crop up.
_NAME_ALIASES = {
    "Rival":       "Terry",
    "Blue":        "Terry",   # calc names the rival "Blue" (vanilla RBY name);
                               # rr_trainers.json stores them as "Terry".
    "Rival Blue":  "Terry",
    # The Team Rocket tab labels the various rocket-grunt fights as
    # "Guard" / "Left Guard" / "Right Guard" (at Game Corner and Rocket
    # Hideout) — in-game these are just Team Rocket Grunts. Aliasing lets
    # the rr_trainers.json matcher pull a free Grunt slot for each one
    # rather than dropping them as unmatched.
    "Guard":       "Grunt",
    "Left Guard":  "Grunt",
    "Right Guard": "Grunt",
}


# Tokens that, when found in the FIRST line of a 2-line header, signal that
# line is a LOCATION rather than a CLASS. E.g. "ROUTE 22 #1\nRIVAL" should
# parse as (area="route_22", name="Rival") not (cls="Route 22", name="Rival").
_LOCATION_TOKENS = (
    "ROUTE", "CITY", "TOWN", "FOREST", "CAVE", "ISL", "ISLAND", "MT.", "MT ",
    "TOWER", "MANSION", "GYM", "DOJO", "LAB", "MUSEUM", "S.S.", "SILPH",
    "ROCKET", "PLATEAU", "POWER PLANT", "VICTORY", "TANOBY", "NAVEL",
    "INDIGO", "PLATEAU", "PATH", "CHAMBER", "RUIN", "ZONE", "VIRID.",
    "FUCHSIA", "CERULEAN", "VERMILION", "VERMILLION", "SAFFRON", "CINNABAR",
    "CELADON", "LAVENDER", "PEWTER", "VIRIDIAN", "PALLET", "FAR.", "BIRTH",
    # Team Rocket / Gauntlet sheet uses these short forms in 2-line
    # headers — without them, Nugg. Bridge Grunt / Dig House Grunt /
    # Rocket Hide. Left Guard / Game Corner Guard / Cerulea. Cave Archer
    # would be mis-parsed as class+name (area "") instead of area+role.
    "BRIDGE", "HIDE.", "GAME CORNER", "DIG HOUSE", "GAUNTLET", "BILL'S",
    "BILLS", "TRAINER TOWER", "BERRY", "POKEMON LEAGUE", "POKÉMON LEAGUE",
    "TUNNEL", "CANYON", "OAK'S", "OAKS", "DOTTED", "LOST", "ICEFALL",
    "ROCK TUNNEL", "ALTERING", "DUNSPARCE", "MAPSE", "FAR.", "ANNE",
    "KINDLE", "BOND", "CAPE", "SEVAULT", "CHRONO", "GREEN PATH",
)


def _looks_like_location(s: str) -> bool:
    su = s.upper().strip()
    # Phrases that are class names containing location-y words (e.g.
    # "GYM LEADER", "ELITE FOUR") — explicitly NOT locations.
    if su in {"GYM LEADER", "ELITE FOUR", "CHAMPION", "RIVAL",
              "ACE TRAINER", "POKEMON TRAINER", "POKÉMON TRAINER",
              "TEAM ROCKET", "ROCKET ADMIN", "ROCKET BOSS"}:
        return False
    return any(tok in su for tok in _LOCATION_TOKENS)


# Roles that, when they appear as the bottom line of a "location\nrole"
# 2-line header, are RR's generic Team Rocket Grunt fight (Nugg. Bridge
# Grunt, Game Corner Guard, Rocket Hide. Left Guard, etc.). For the
# calc_label / class we expand them all to "Team Rocket Grunt" so they
# share a clean canonical name, and the LOCATION carries through as the
# fight_label that disambiguates which encounter this is.
_GRUNT_ROLE_NAMES = {"GRUNT", "GUARD", "LEFT GUARD", "RIGHT GUARD"}


def _parse_header(cell_value) -> tuple[str, str, str, str] | None:
    """Parse a trainer header cell.

    Headers come in three shapes:
      2-line + class first:    "GYM LEADER\\nBROCK"
          → (cls="Gym Leader",  name="Brock",   area="",
             header_location="")
      2-line + location first: "NUGG. BRIDGE\\nGRUNT"
          → (cls="Team Rocket Grunt", name="Grunt", area="nugget_bridge",
             header_location="Nugg. Bridge")
      3-line:                  "PEWTER MUSEUM\\nLEADER\\nFALKNER"
          → (cls="Leader",      name="Falkner", area="pewter_museum",
             header_location="")

    The fourth element (header_location) is the raw display location for
    2-line location-first headers — used by parse_boss_sheet to set
    fight_label so calc_label includes the location ("Team Rocket Grunt
    (Nugg. Bridge)"), keeping every per-area grunt fight distinct.

    Returns None if the cell isn't a trainer header.
    """
    s = _norm_str(cell_value)
    if not s or "\n" not in s:
        return None
    lines = [ln.strip() for ln in s.split("\n") if ln.strip()]
    if len(lines) < 2:
        return None
    header_location = ""
    if len(lines) == 2:
        first, second = lines
        if _looks_like_location(first):
            area = _normalize_area_name(first)
            header_location = first.title()
            # "GRUNT" / "GUARD" / "LEFT GUARD" / "RIGHT GUARD" → all collapse
            # to the canonical "Team Rocket Grunt" class. Other roles
            # (rivals, "RIVAL" on "ROUTE 22 #1\\nRIVAL", etc.) reuse the
            # bottom line as both class and name.
            second_upper = second.upper()
            if second_upper in _GRUNT_ROLE_NAMES:
                cls = "Team Rocket Grunt"
                nm  = "Grunt"
            else:
                cls = second
                nm  = second
        else:
            # (class, name) — standard 2-line.
            cls  = first
            nm   = second
            area = ""
    else:
        # 3-line: location, class, name. Most reliable shape.
        area_disp = lines[0]
        cls       = " ".join(lines[1:-1])
        nm        = lines[-1]
        area      = _normalize_area_name(area_disp)
    cls = cls.strip().title()
    nm  = nm.strip().title()
    if not cls or not nm:
        return None
    if nm.upper() in {"BASE STATS", "EVS", "IVS", "MOVES", "SPEED STAT:"}:
        return None
    if cls.lower() in {"section", "page", "navigation", "table of contents"}:
        return None
    return (cls, nm, area, header_location)


def _parse_level(raw) -> tuple[int, int | None]:
    """Parse a level cell into (level, offset).

    The spreadsheet encodes mon levels in three forms:
      • A plain number (e.g. "22.0")  →  fixed level 22
      • "Max Level"                    →  player's highest mon at fight time
      • "Max Level - N" / "Highest Lv - N"  →  highest minus N (e.g. -2)

    Relative levels return (0, offset_int) where offset_int <= 0; the dashboard
    resolves them at render time against the player's current highest party
    level. Fixed levels return (level, None). Unknown levels return (0, None).
    """
    if raw is None:
        return (0, None)
    if isinstance(raw, (int, float)):
        return (int(raw), None)
    s = str(raw).strip()
    if not s:
        return (0, None)
    # "Max Level" / "Max Level - 2" / "Highest Lv - 2" / "Highest Lv -2" / etc.
    m = re.search(
        r"(?:Max\s*Level|Highest\s*Lv)(?:\s*[-–]\s*(\d+))?",
        s, re.I,
    )
    if m:
        offset = int(m.group(1) or 0)
        return (0, -offset)   # 0 → "Max Level", -2 → "Max Level - 2"
    try:
        return (int(float(s)), None)
    except ValueError:
        return (0, None)


def _parse_mon_block(ws, header_row: int, col: int) -> dict | None:
    """Extract one mon at (header_row, col). Returns None if empty slot."""
    name = _norm_str(ws.cell(row=header_row, column=col).value)
    if not name:
        return None
    # Skip cells that are actually labels rather than species (e.g. "BASE STATS").
    if name.isupper() and len(name) > 15:
        return None
    if name.upper() in {"BASE STATS", "EVS", "IVS", "MOVES", "SPEED STAT:"}:
        return None
    level_raw = ws.cell(row=header_row + _OFF_LEVEL,   column=col).value
    nature    = _norm_str(ws.cell(row=header_row + _OFF_NATURE,  column=col).value)
    ability   = _norm_str(ws.cell(row=header_row + _OFF_ABILITY, column=col).value)
    item      = _norm_str(ws.cell(row=header_row + _OFF_ITEM,    column=col).value)
    moves: list[str] = []
    for i in range(4):
        mv = _norm_str(ws.cell(row=header_row + _OFF_MOVE_0 + i, column=col).value)
        if mv:
            moves.append(_normalize_move(mv))
    level, level_offset = _parse_level(level_raw)
    entry = {
        "species": name,
        "level":   level,
        "nature":  nature,
        "ability": ability,
        "item":    item,
        "moves":   moves,
    }
    if level_offset is not None:
        # "Max Level" / "Max Level - N" / "Highest Lv - N" — render-time
        # resolution against the player's highest party mon level.
        entry["level_offset"] = level_offset
    return entry


def _humanize_section_label(raw: str) -> str:
    """Turn "(!) PRE LT. SURGE" into "Pre-Lt. Surge" — the spreadsheet author's
    intent is the user-facing string, but the SHOUTING-CAPS form is jarring.
    """
    s = raw.strip()
    if s.startswith("(!)"):
        s = s[3:].strip()
    # Title-case, but preserve a few well-known idioms.
    fixed = s.title()
    # Re-cap "Lt." (Title would lower the period), "PKMN", "&".
    fixed = re.sub(r"\bLt\.\s*", "Lt. ", fixed)
    return fixed


# Move-name normalisation: the spreadsheet truncates long move names with
# trailing "." (e.g. "First Impress." → "First Impression") so they fit in
# the narrow column. The calc's pokedex keys use the full canonical names,
# so abbreviated names fail the Prep-tab moveset lookup and render as
# blanks. Map known abbreviations to their full names and apply at parse
# time so both the JSON output AND the supplementary calc sets carry the
# normalised form.
_MOVE_NAME_MAP: dict[str, str] = {
    # Sheet abbreviation → canonical calc name.
    "Behemo. Blade":   "Behemoth Blade",
    "Behemo. Bash":    "Behemoth Bash",
    "Clang. Scales":   "Clanging Scales",
    "Clang. Soul":     "Clangorous Soul",
    "Dazz. Gleam":     "Dazzling Gleam",
    "Disarm. Voice":   "Disarming Voice",
    "Double I. Bash":  "Double Iron Bash",
    "Dragon Hamm.":    "Dragon Hammer",
    "Dyna. Cannon":    "Dynamax Cannon",
    "Dynam. Punch":    "Dynamic Punch",
    "Expand. Force":   "Expanding Force",
    "First Impress.":  "First Impression",
    "Giga. Hammer":    "Gigaton Hammer",
    "Headlon. Rush":   "Headlong Rush",
    "High Horsep.":    "High Horsepower",
    "High J. Kick":    "High Jump Kick",
    "Hypersp. Fury":   "Hyperspace Fury",
    "Hypersp. Hole":   "Hyperspace Hole",
    "Matcha Got.":     "Matcha Gotcha",
    "Parab. Charge":   "Parabolic Charge",
    "Popula. Bomb":    "Population Bomb",
    "Pow-Up Punch":    "Power-Up Punch",
    "Scorch. Sands":   "Scorching Sands",
    "Steam Erupt.":    "Steam Eruption",
    "Stom. Tantrum":   "Stomping Tantrum",
    "Surg. Strikes":   "Surging Strikes",
    "Water Shurik.":   "Water Shuriken",
    "Vacuum W.":       "Vacuum Wave",
    "Crush Grip":      "Crush Grip",
    # Case / hyphenation differences that the calc is picky about.
    "Roar Of Time":    "Roar of Time",
    "Soft Boiled":     "Soft-Boiled",
    "U-Turn":          "U-turn",
    "V-Create":        "V-create",
    "X-Scissor":       "X-Scissor",
    # Hidden Power: sheet abbreviates as "HP <Type>", calc uses
    # "Hidden Power <Type>" capitalised.
    "HP Dark":         "Hidden Power Dark",
    "HP Dragon":       "Hidden Power Dragon",
    "HP Electric":     "Hidden Power Electric",
    "HP Fighting":     "Hidden Power Fighting",
    "HP Fire":         "Hidden Power Fire",
    "HP Flying":       "Hidden Power Flying",
    "HP Ghost":        "Hidden Power Ghost",
    "HP Grass":        "Hidden Power Grass",
    "HP Ground":       "Hidden Power Ground",
    "HP Ice":          "Hidden Power Ice",
    "HP Poison":       "Hidden Power Poison",
    "HP Psychic":      "Hidden Power Psychic",
    "HP Rock":         "Hidden Power Rock",
    "HP Steel":        "Hidden Power Steel",
    "HP Water":        "Hidden Power Water",
}


def _normalize_move(name: str) -> str:
    """Resolve a sheet move name to its calc-canonical form.

    Falls through unchanged when no mapping is known — moves like "Absorb",
    "Pound", "Tail Whip" are valid but absent from normal.js because they
    don't appear in any RR trainer set; the calc still accepts them, it
    just has no precomputed data.
    """
    if not name:
        return name
    s = name.strip()
    if s in _MOVE_NAME_MAP:
        return _MOVE_NAME_MAP[s]
    # Smart fallback: drop a trailing "." and try Title-Case match.
    if s.endswith("."):
        trimmed = s[:-1].strip()
        if trimmed in _MOVE_NAME_MAP:
            return _MOVE_NAME_MAP[trimmed]
    return s


_SPRITE_URL_RE = re.compile(r'IMAGE\("([^"]+)"', re.I)


def _extract_image_url(formula_cell_value) -> str:
    """Pull the URL out of `=IMAGE("https://...")`-style cell formulas."""
    if not isinstance(formula_cell_value, str):
        return ""
    m = _SPRITE_URL_RE.search(formula_cell_value)
    return m.group(1) if m else ""


def parse_boss_sheet(ws, ws_formulas=None) -> list[dict]:
    """Walk a boss-format sheet and yield trainer-block entries.

    Each entry has: class, name, area (from 3-line header, may be ""),
    party, fight_label, source.

    fight_label captures the spreadsheet's human-readable "when does this
    variant trigger" annotation. We pull it from two places, preferring the
    more meaningful one:

      • Col 5 "(!) X" section banners (e.g. "(!) PRE LT. SURGE",
        "(!) IF RIVAL HAS SQUIRTLE", "(!) TEAM ONE") — the canonical,
        user-facing label the sheet author wrote.
      • Col 3 "IF YOU'RE LEVEL N ->>" two-row condition (e.g. "LEVEL 27 ->>")
        — the level-based variant flag. Only used as a fallback when no
        "(!) ..." banner is in effect.

    A section banner stays in effect until the next "(!) ..." banner — so
    one banner ("(!) IF RIVAL HAS SQUIRTLE") can label multiple consecutive
    trainer headers (the same variant across multiple rival fights).
    """
    out: list[dict] = []
    pending_level_condition: str = ""
    current_section: str = ""   # from "(!) X" col-5 banners; sticks across headers
    r = 1
    while r <= ws.max_row:
        # Scan the action columns (5/10/15) for a "(!) X" section banner —
        # update the running section label whenever one appears.
        for c in (5, 10, 15):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str) and v.strip().startswith("(!)"):
                current_section = _humanize_section_label(v)
                break

        cell = ws.cell(row=r, column=_HEADER_COL).value
        s = _norm_str(cell)
        # Two-row "IF YOU'RE\nLEVEL N ->>" condition tag (col 3) — feeds the
        # fallback level-based label.
        if s.upper().startswith("IF YOU"):
            next_v = _norm_str(ws.cell(row=r + 1, column=_HEADER_COL).value)
            m = re.search(r"LEVEL\s+(\d+)", next_v.upper())
            if m:
                pending_level_condition = f"If Lv ≥ {m.group(1)}"
            r += 2
            continue
        header = _parse_header(cell)
        if not header:
            r += 1
            continue
        cls, name, area, header_location = header
        party: list[dict] = []
        for c in _MON_COLS:
            mon = _parse_mon_block(ws, r, c)
            if mon:
                party.append(mon)
        if not party:
            r += 1
            continue
        # Banner wins. Then location-from-header (for generic-role 2-line
        # headers like "Nugg. Bridge\\nGrunt", where the location IS the
        # disambiguator between otherwise-identical Team-Rocket-Grunt
        # entries). Then "IF YOU'RE LEVEL N" fallback.
        fight_label = (current_section
                       or header_location
                       or pending_level_condition
                       or "")
        # Capture the trainer sprite URL from the row ABOVE the text header
        # — the spreadsheet places `=IMAGE("https://i.ibb.co/.../Name.png")`
        # one row up from the text label in col 3.
        sprite_url = ""
        if ws_formulas is not None and r >= 2:
            sprite_url = _extract_image_url(
                ws_formulas.cell(row=r - 1, column=_HEADER_COL).value)
        out.append({
            "class":       cls,
            "name":        name,
            "area":        area,
            "party":       party,
            "fight_label": fight_label,
            "sprite_url":  sprite_url,
            "source":      f"{ws.title}!R{r}",
        })
        pending_level_condition = ""
        r += 1
    return out


# ── Area name → area_id normalization ──────────────────────────────────────────
# The Trainer Order sheet uses display names like "PEWTER CITY" / "ROUTE 25".
# data/games/gen3_frlge/area_map.json uses snake_case IDs like "pewter_city"
# and "route_25". This map handles the common patterns; unknown areas pass
# through the lowercase+underscore normalization.

_AREA_OVERRIDES = {
    # Areas where the display name doesn't map 1:1 to an area_id snake_case.
    "S.S. ANNE":            "ss_anne",
    "S.S. AQUA":            "ss_aqua",
    "POKÉMON LEAGUE":       "indigo_plateau",
    "POKEMON LEAGUE":       "indigo_plateau",
    "INDIGO PLATEAU":       "indigo_plateau",
    "MT. MOON":             "mt_moon",
    "MT. EMBER":            "mt_ember",
    "MT. SILVER":           "mt_silver",
    "POKÉMON TOWER":        "pokemon_tower",
    "POKEMON TOWER":        "pokemon_tower",
    "PKMN TOWER":           "pokemon_tower",
    "POKÉMON MANSION":      "pokemon_mansion",
    "POKEMON MANSION":      "pokemon_mansion",
    "POKÉMON MANSION ENTRANCE": "pokemon_mansion",
    "POKEMON MANSION ENTRANCE": "pokemon_mansion",
    "ROCKET HIDEOUT":       "rocket_hideout",
    "ROCKET WAREHOUSE":     "rocket_warehouse",
    "DIGLETT'S CAVE":       "digletts_cave",
    "DIGLETTS CAVE":        "digletts_cave",
    "VIRIDIAN FOREST":      "viridian_forest",
    "VIRID. FOREST":        "viridian_forest",
    "VIRDIAN FOREST":       "viridian_forest",   # Nuzlocke Redux typo
    "VICTORY ROAD":         "victory_road",
    "POWER PLANT":          "power_plant",
    "BERRY FOREST":         "berry_forest",
    "ICEFALL CAVE":         "icefall_cave",
    "ROCK TUNNEL":          "rock_tunnel",
    "CERULEAN CAVE":        "cerulean_cave",
    "CERULEA CAVE":         "cerulean_cave",     # spreadsheet abbreviation
    "SEAFOAM ISLANDS":      "seafoam_islands",
    "SEAFOAM ISL.":         "seafoam_islands",
    "SAFARI ZONE":          "safari_zone_center",
    "DOTTED HOLE":          "dotted_hole",
    "LOST CAVE":            "lost_cave",
    "NAVEL ROCK":           "navel_rock",
    "BIRTH ISLAND":         "birth_island",
    "FARAWAY ISLAND":       "faraway_island",
    "FAR. ISLAND":          "faraway_island",
    "TANOBY RUINS":         "tanoby_ruins",
    "SILPH CO.":            "silph_co",
    "SILPH CO":             "silph_co",
    "TRAINER TOWER":        "trainer_tower",
    "NUGGET BRIDGE":        "nugget_bridge",
    "NUGG. BRIDGE":         "nugget_bridge",
    "BOND BRIDGE":          "bond_bridge",
    "KINDLE ROAD":          "kindle_road",
    "CAPE BRINK":           "cape_brink",
    "SEVAULT CANYON":       "sevault_canyon",
    "CHRONO ISLAND":        "chrono_island",
    "GREEN PATH":           "green_path",
    "CANYON ENTRANCE":      "sevault_canyon",
    "OAK'S LABORATORY":     "oaks_lab",
    "OAKS LABORATORY":      "oaks_lab",
    "OAK'S LAB":            "oaks_lab",
    # "X City Gym" → just the city.
    "PEWTER CITY GYM":      "pewter_city",
    "CERULEAN CITY GYM":    "cerulean_city",
    "VERMILION CITY GYM":   "vermilion_city",
    "VERMILLION CITY GYM":  "vermilion_city",
    "CELADON CITY GYM":     "celadon_city",
    "FUCHSIA CITY GYM":     "fuchsia_city",
    "FUSCHIA CITY GYM":     "fuchsia_city",      # Nuzlocke Redux typo
    "SAFFRON CITY GYM":     "saffron_city",
    "CINNABAR ISLAND GYM":  "cinnabar_island",
    "VIRIDIAN CITY GYM":    "viridian_city",
    "CELADON CITY GAME CORNER": "celadon_city_game_corner",
    "GAME CORNER":             "celadon_city_game_corner",
    "ROCKET HIDE.":            "rocket_hideout",
    "ROCKET HIDE":             "rocket_hideout",
    "CERULEA. CAVE":           "cerulean_cave",
    "CERULEAN CAVE":           "cerulean_cave",
    "CHAMPION":             "indigo_plateau",
    "ELITE FOUR":           "indigo_plateau",
    # RR has multiple "Route 22" encounters (early + late game). Both map
    # to the single area_id reported by the game ("route_22"); the renderer
    # uses level_cap proximity to distinguish them into separate rows.
    "ROUTE 22 #1":          "route_22",
    "ROUTE 22 #2":          "route_22",
    "VIRID. FOREST #1":     "viridian_forest",
    "VIRID. FOREST #2":     "viridian_forest",
}


def _normalize_area_name(display: str) -> str:
    if not display:
        return ""
    s = display.strip()
    if s.upper() in _AREA_OVERRIDES:
        return _AREA_OVERRIDES[s.upper()]
    # Lowercase + collapse spaces to underscores + strip punctuation
    s = s.lower()
    s = re.sub(r"[.''']", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s


# Class words to strip from calc trainer labels when extracting the bare
# trainer name. Calc uses tags like "Leader Brock", "Hiker Franklin",
# "Bug Catcher Robby" — we want just "Brock", "Franklin", "Robby" for
# rr_trainers.json matching.
_CALC_CLASS_PREFIXES = (
    # Multi-word first (longest match wins)
    "PkMn Trainer", "PkMn Ranger", "Ace Trainer", "Bug Catcher",
    "Black Belt", "Bird Keeper", "Battle Girl", "Cooltrainer",
    "Cue Ball", "Crush Girl", "Crush Kin", "Elite Four",
    "Parasol Lady", "Rocket Admin", "Rocket Boss", "Rocket Executive",
    "Ruin Maniac", "Super Nerd", "Team Rocket Grunt", "Team Rocket",
    "Young Couple", "Sis & Bro", "Aroma Lady", "Dumbass",
    # Single-word
    "Leader", "Champion", "Rival", "Picnicker", "SwimmerM", "SwimmerF",
    "Channeler", "Biker", "Hiker", "Lass", "Scientist", "Youngster",
    "Fisherman", "Camper", "Pokemaniac", "Juggler", "Gambler", "Sailor",
    "Gentleman", "Beauty", "Burglar", "Painter", "Tamer", "Rocker",
    "Psychic", "Twins", "Lady", "Pokefan", "Maniac", "Boss", "Admin",
    "Goon", "Ranger", "Engineer", "Coach", "Teacher", "Professor",
)


def _strip_calc_class(label: str) -> tuple[str, str]:
    """Given "Leader Brock" → ("Leader", "Brock"). "*Leader Bugsy Set 1" →
    ("*Leader", "Bugsy"). Returns (class, bare_name). Multi-word class
    prefixes are matched in order (longest first).
    """
    cleaned = label.strip()
    # Capture/strip leading asterisk marker (post-LT-Surge variant flag)
    asterisk = ""
    if cleaned.startswith("*"):
        asterisk = "*"
        cleaned = cleaned[1:].lstrip()
    # Strip trailing " Set N" / " Variant N" sub-label.
    cleaned = re.sub(r"\s+(Set|Variant)\s+\d+\s*$", "", cleaned)
    for prefix in _CALC_CLASS_PREFIXES:
        if cleaned.startswith(prefix + " "):
            cls = (asterisk + prefix).strip()
            name = cleaned[len(prefix) + 1:].strip()
            return (cls, name)
    # Fallback: split on first space
    parts = cleaned.split(None, 1)
    if len(parts) == 2:
        return (asterisk + parts[0], parts[1])
    return (asterisk, cleaned)


def parse_calc_sets(path: Path) -> list[dict]:
    """Parse calc/src/js/data/sets/normal.js into trainer-party blocks.

    Inverts the species→trainer dict so each output entry is one trainer's
    full party. Returns list of {class, name, party, source} entries with
    the same shape as parse_boss_sheet.
    """
    if not path.exists():
        return []
    txt = path.read_text(encoding="utf-8")
    m = re.match(r"\s*var\s+SETDEX_SV\s*=\s*(\{.+\});\s*$", txt, re.S)
    if not m:
        return []
    data = json.loads(m.group(1))
    # Invert: trainer_label → [mon, ...]
    raw: dict[str, list[dict]] = defaultdict(list)
    for species, sets in data.items():
        for set_name, set_data in sets.items():
            raw[set_name].append({
                "species": species,
                "level":   int(set_data.get("level", 0)),
                "nature":  set_data.get("nature", ""),
                "ability": set_data.get("ability", ""),
                "item":    set_data.get("item", ""),
                "moves":   list(set_data.get("moves", [])),
                "evs":     set_data.get("evs", {}),
                "ivs":     set_data.get("ivs", {}),
            })
    out: list[dict] = []
    for label, mons in raw.items():
        cls, name = _strip_calc_class(label)
        # If the label carries a "Set N" suffix, capture it as a fight_label
        # so the dashboard can disambiguate multiple fights against the same
        # trainer (Bugsy has different teams depending on when you fight him,
        # encoded in calc as "Leader Bugsy", "Leader Bugsy Set 1", "Leader
        # Bugsy Set 2", …).
        # We intentionally do NOT use calc's "Set N" suffix as fight_label —
        # it's opaque to users. The xlsx sheet's "IF YOU'RE LEVEL N" condition
        # is the canonical, human-readable label and is merged in later from
        # sheet_fight_label_by_key.
        fight_label = ""
        # Sort party by lead level ascending — keeps Pokemon order stable
        # across reruns. (Original order in the JS is alphabetical by species,
        # which doesn't reflect in-game lead order — but for the dashboard
        # we want a deterministic, sensible ordering.)
        mons.sort(key=lambda m: (m["level"], m["species"]))
        out.append({
            "class":       cls.title(),
            "name":        name.title(),
            "area":        "",
            "party":       mons,
            "calc_label":  label,         # raw label, used by Prep tab handoff
            "fight_label": fight_label,   # "Set 1", "Variant 2", or ""
            "source":      f"calc:{label}",
        })
    return out


def parse_main_milestones(ws) -> tuple[list[tuple[str, int]], dict[str, int], dict[str, int]]:
    """Parse the Main sheet's "Pre-X (cap)" milestone list into structured caps.

    The Main tab lists the level cap at each story checkpoint, e.g.:
        Pre-Brock (15)
        Pre-Mt. Moon Archer (22)
        Pre-Misty (27)
        Pre-Lt. Surge (34)
        ...
        Post Game (100)

    Returns three views:
      • order: [(milestone_name, cap), …] in story order
      • pre:   {milestone_name: cap}             — "Pre X" cap (fight against X)
      • post:  {milestone_name: cap_of_next}     — "Post X" cap (after beating X)
                                                    falls back to 100 at end of list
    """
    order: list[tuple[str, int]] = []
    label_re = re.compile(r"Pre-?\s*(.+?)\s*\((\d+)\)", re.I)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            m = label_re.search(v)
            if m:
                order.append((m.group(1).strip(), int(m.group(2))))
                break  # one milestone per row
        # Also catch the "Post Game (100)" suffix.
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "Post Game" in v:
                m2 = re.search(r"\((\d+)\)", v)
                if m2 and not any(n == "Post Game" for n, _ in order):
                    order.append(("Post Game", int(m2.group(1))))
                break
    pre  = {name: cap for name, cap in order}
    post: dict[str, int] = {}
    for i, (name, _cap) in enumerate(order):
        post[name] = order[i + 1][1] if i + 1 < len(order) else 100
    return order, pre, post


def parse_trainer_order(ws) -> tuple[list[tuple[str, str, bool]],
                                      dict[str, list[int]],
                                      dict[tuple[str, str], list[int]]]:
    """Parse Trainer Order sheet.

    Each block is 2 rows: row N has trainer name + level cap, row N+1 has
    area name. "(OPTIONAL)" markers in col 3 flag the next trainer as
    skippable. Returns three views over the data:
      - pairs:           [(trainer_name_title, area_id, is_optional), …]
      - caps_by_name:    {name: [cap, …]} (encounter-order list)
      - caps_by_name_area: {(name, area_id): [cap, …]}
        Same trainer can fight in the same area multiple times (e.g. the
        Rival appears at Route 22 twice in RR), so each value is a list.
    """
    pairs: list[tuple[str, str, bool]] = []
    caps_by_name: dict[str, list[int]] = defaultdict(list)
    caps_by_name_area: dict[tuple[str, str], list[int]] = defaultdict(list)
    pending_optional = False
    last_trainer: str | None = None
    last_cap: int | None = None
    for r in range(1, ws.max_row + 1):
        col3 = _norm_str(ws.cell(row=r, column=3).value)
        col4 = _norm_str(ws.cell(row=r, column=4).value)
        if col3 == "(OPTIONAL)":
            pending_optional = True
            continue
        if not col4:
            continue
        level_cap = ws.cell(row=r, column=6).value
        if level_cap is not None and last_trainer is None:
            last_trainer = col4
            try:
                last_cap = int(float(level_cap))
            except (TypeError, ValueError):
                last_cap = None
        elif last_trainer is not None:
            area_id = _normalize_area_name(col4)
            nm_t = last_trainer.title()
            pairs.append((nm_t, area_id, pending_optional))
            if last_cap is not None:
                caps_by_name[nm_t].append(last_cap)
                caps_by_name_area[(nm_t, area_id)].append(last_cap)
            last_trainer = None
            last_cap = None
            pending_optional = False
        else:
            last_trainer = col4
            last_cap = None
    return pairs, dict(caps_by_name), dict(caps_by_name_area)


def _build_name_index(rr_trainers_path: Path) -> dict[str, list[dict]]:
    """Build {trainer_name → [{rt_id, party_size, class_id}, ...]} from rr_trainers.json.

    rr_trainers.json stores 0-based array indices as keys; runtime IDs are
    1-based, so we add 1.  Names are not unique (e.g. multiple "Coach", "Brendan",
    "Falkner" entries — each is a distinct fight, possibly a rematch). The
    consumer picks the best unused candidate.
    """
    with rr_trainers_path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    idx: dict[str, list[dict]] = defaultdict(list)
    for k, v in data.get("trainers", {}).items():
        nm = (v.get("name") or "").strip()
        if not nm:
            continue
        runtime_id = int(k) + 1  # 0-based → 1-based
        idx[nm.title()].append({
            "rt_id":      runtime_id,
            "party_size": int(v.get("party_size", 0)),
            "class_id":   int(v.get("class", 0)),
        })
    return dict(idx)


def _pick_trainer_id(name_index: dict[str, list[dict]],
                     name: str, party_size: int,
                     used_ids: set[int]) -> int | None:
    """Pick the best unused runtime trainer ID for a parsed trainer.

    Strategy:
      1. Look up candidates by name (Title case) in rr_trainers.json index.
      2. Filter out already-used IDs (a parsed trainer block claims one ID,
         so subsequent same-name blocks pick the next one).
      3. Prefer candidates where rr_trainers.json party_size matches the
         parsed party_size exactly; otherwise pick the closest party_size,
         otherwise the lowest rt_id.
    Returns None when no name match exists.
    """
    candidates = name_index.get(name.title()) or []
    available = [c for c in candidates if c["rt_id"] not in used_ids]
    if not available:
        # All same-name slots taken — fall back to any candidate (first match)
        # so we still surface the party data even if the rt_id is reused.
        return candidates[0]["rt_id"] if candidates else None
    # Score: exact party_size match wins; otherwise closest, then lowest rt_id.
    available.sort(key=lambda c: (
        abs(c["party_size"] - party_size),
        c["rt_id"],
    ))
    return available[0]["rt_id"]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--src",
                        default=str(_REPO_ROOT / "rr_trainers_dump.xlsx"),
                        help="Path to the downloaded xlsx (default: repo root)")
    parser.add_argument("--out", default=str(_OUT_PATH),
                        help=f"Output path (default: {_OUT_PATH})")
    args = parser.parse_args()

    src = Path(args.src)
    if not src.exists():
        print(f"Source xlsx not found: {src}", file=sys.stderr)
        return 1

    print(f"Loading {src}…")
    # Two views: data_only=True evaluates formulas (so we read computed
    # level/move values); the formula-form view exposes raw `=IMAGE("url")`
    # cells so we can pull trainer sprite URLs.
    wb = load_workbook(str(src), data_only=True)
    wb_formulas = load_workbook(str(src), data_only=False)

    # 1a) Parse calc/src/js/data/sets/normal.js — PRIMARY party source.
    # The damage calculator already consumes this data, so it's the canonical
    # set definition used at runtime. Provides nature/ability/item/moves/IVs/EVs
    # for every trainer.
    #
    # Filter: drop entries with > 6 mons. A real trainer party is capped at 6
    # — anything larger is a calc-side SUPERSET (e.g. "*Leader Misty" with 11
    # mons that the game narrows to 6 at fight time based on starter choice).
    # Asterisk-prefixed labels are NOT inherently bad: small ones (like
    # "*Leader Bugsy Set 1" with 2 mons) are legit fight variants and we want
    # to keep them so the dashboard surfaces a calc_label for those fights.
    calc_all = parse_calc_sets(_CALC_NORMAL_PATH)
    calc_trainers = [t for t in calc_all if len(t.get("party", [])) <= 6]
    skipped_supersets = len(calc_all) - len(calc_trainers)
    print(f"  calc/normal.js : {len(calc_trainers)} trainers parsed "
          f"(canonical sets — primary party source; "
          f"skipped {skipped_supersets} >6-mon supersets)")

    # 1b) Parse trainer parties from every boss sheet — used for AREA mapping
    # (3-line headers like "PEWTER MUSEUM\\nLEADER\\nFALKNER") and as a
    # supplemental source for trainers the calc doesn't cover.
    sheet_trainers: list[dict] = []
    for sheet in _BOSS_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"  skip — missing sheet {sheet!r}")
            continue
        entries = parse_boss_sheet(wb[sheet], wb_formulas[sheet])
        print(f"  {sheet:<18} {len(entries):>4} trainers parsed")
        sheet_trainers.extend(entries)

    # Build quick lookups from sheet entries — used to enrich calc-sourced
    # trainers with the sheet's area (from 3-line headers) and fight_label
    # (the human-readable "(!) X" banner the sheet author wrote — e.g.
    # "Pre Lt. Surge", "If Rival Has Squirtle").
    sheet_area_by_name: dict[str, list[str]] = defaultdict(list)
    # Exact-party-size lookup (sheet "Bugsy" with 4 mons matches calc "Bugsy"
    # only if calc also has a 4-mon Bugsy). Stored as a list per key — when
    # multiple sheet entries share the same (name, party_size) the label is
    # AMBIGUOUS (e.g. several Grunt fights at different locations all have
    # party_size 4) and the merge step refuses to attach it to calc copies.
    sheet_fight_label_by_key: dict[tuple[str, int], list[str]] = defaultdict(list)
    # Ordered fallback: when exact size doesn't match, use the calc's
    # "Set N" position to index into the sheet's chronological list of
    # labels for that trainer name. Sheet rows are in story order, so
    # labels[0] = first fight, labels[1] = second fight, etc.
    sheet_fight_labels_by_name: dict[str, list[str]] = defaultdict(list)
    for s in sheet_trainers:
        if s.get("area"):
            sheet_area_by_name[s["name"].title()].append(s["area"])
        if s.get("fight_label"):
            key = (s["name"].title(), len(s.get("party", [])))
            sheet_fight_label_by_key[key].append(s["fight_label"])
            sheet_fight_labels_by_name[s["name"].title()].append(
                s["fight_label"])

    def _extract_set_index(calc_label: str) -> int:
        """Return the 1-based Set/Variant number in a calc label, or 1.

        '*Leader Bugsy Set 2' → 2, 'Leader Bugsy Set 1' → 1, 'Leader Bugsy' → 1.
        """
        m = re.search(r"\s+(?:Set|Variant)\s+(\d+)\s*$", calc_label or "")
        return int(m.group(1)) if m else 1

    # Merge: SHEET IS CANONICAL for the data it covers, but the calc
    # carries many more distinct FIGHTS than the sheet enumerates —
    # especially repeated trainer classes like "Team Rocket Grunt Set 1..7"
    # where each Set is a different grunt at a different location. Dropping
    # those calc entries because the sheet has a 2-mon Grunt with the
    # matching party_size would hide whole encounters from the widget.
    #
    # Dedup key is the CALC_LABEL (synthesised "Class Name (Fight)" for
    # sheet entries; original "Class Name Set N" for calc entries). Sheet
    # and calc labels almost never collide, so all sheet entries and all
    # calc entries flow through. Within calc, the label IS unique by
    # construction. Within sheet, our synthesis combines class+name+
    # fight_label which is unique per fight.
    raw_trainers: list[dict] = []
    seen_calc_labels: set[str] = set()
    for s in sheet_trainers:
        for mon in s.get("party") or []:
            mon["species"] = _normalize_species(mon.get("species") or "")
        s["calc_label"] = _synthesise_calc_label(
            s.get("class") or "", s["name"], s.get("fight_label") or "")
        raw_trainers.append(s)
        seen_calc_labels.add(s["calc_label"])

    calc_added = 0
    for t in calc_trainers:
        if t.get("calc_label") in seen_calc_labels:
            continue
        name_t = t["name"].title()
        # Only inherit a sheet area for the calc copy when the sheet's
        # mention of this name is UNAMBIGUOUS — a single location. When the
        # name shows up at multiple locations (e.g. Grunt appears at both
        # Nugg. Bridge and Dig House), guessing one would silently glue
        # every calc copy to that one area, which is exactly the "merging
        # trainers that shouldn't be merged" problem. Leave area blank;
        # the trainers_by_area pass distributes calc copies via the
        # one-per-name queue.
        areas = sheet_area_by_name.get(name_t) or []
        if len(set(areas)) == 1:
            t["area"] = areas[0]
        # Only attach a fight_label from sheet when it's UNAMBIGUOUS: exactly
        # one sheet entry shares this (name, party_size). Otherwise the calc
        # copy could end up at a different area than the label implies, which
        # is the "wrong fight_label on wrong area" merge bug.
        matches = sheet_fight_label_by_key.get(
            (name_t, len(t.get("party", [])))) or []
        sheet_fl = matches[0] if len(set(matches)) == 1 else ""
        if not sheet_fl:
            labels = sheet_fight_labels_by_name.get(name_t) or []
            # Only use the Set-index fallback when the sheet has the SAME
            # number of distinct fights as calc has Sets — otherwise the
            # index correspondence is unreliable.
            set_idx = _extract_set_index(t.get("calc_label", "")) - 1
            if len(labels) >= 2 and 0 <= set_idx < len(labels):
                sheet_fl = labels[set_idx]
        if sheet_fl:
            t["fight_label"] = sheet_fl
        raw_trainers.append(t)
        seen_calc_labels.add(t["calc_label"])
        calc_added += 1
    print(f"Total trainers : sheet={len(sheet_trainers)} "
          f"+ calc-extra={calc_added} = {len(raw_trainers)}")

    # 2a) Parse the Main sheet's level-cap milestone progression. The output
    # JSON gets a `milestones` block so the dashboard can map a player's
    # current state to "where are they in the story" — and so the
    # _trainer_panel_html renderer can flag each fight variant as Past /
    # Current / Future against the player's highest party level.
    main_milestones: list[tuple[str, int]] = []
    main_pre_caps: dict[str, int] = {}
    main_post_caps: dict[str, int] = {}
    if "Main" in wb.sheetnames:
        main_milestones, main_pre_caps, main_post_caps = parse_main_milestones(
            wb["Main"])
        print(f"Main milestones: {len(main_milestones)} entries "
              f"(first: {main_milestones[0] if main_milestones else None}, "
              f"last: {main_milestones[-1] if main_milestones else None})")

    # 2b) Parse Trainer Order sheet for area mapping AND level caps.
    order_pairs: list[tuple[str, str, bool]] = []
    level_caps: dict[str, list[int]] = {}
    level_caps_by_area: dict[tuple[str, str], list[int]] = {}
    if "Trainer Order" in wb.sheetnames:
        order_pairs, level_caps, level_caps_by_area = parse_trainer_order(
            wb["Trainer Order"])
        print(f"Trainer Order: {len(order_pairs)} (trainer, area) pairs, "
              f"{sum(len(v) for v in level_caps.values())} level caps "
              f"across {len(level_caps)} trainer names "
              f"({len(level_caps_by_area)} (name, area) keys)")

    # 3) Build name → runtime_id index from rr_trainers.json.
    if not _RR_TRAINERS_PATH.exists():
        print(f"Missing {_RR_TRAINERS_PATH}", file=sys.stderr)
        return 1
    name_index = _build_name_index(_RR_TRAINERS_PATH)
    print(f"rr_trainers.json: {len(name_index)} unique trainer names")

    # 4) (Optional) parse Nuzlocke Redux SSR'd HTML for additional area hints.
    redux_locs: dict[str, list[str]] = defaultdict(list)
    redux_html_path = _REPO_ROOT / "redux_page.html"
    if redux_html_path.exists():
        nav_re = re.compile(
            r">([A-Z][A-Za-z'.\- ]*?)\s+at\s+([A-Z][A-Za-z'.,\- ]+?)<",
            re.S,
        )
        html_blob = redux_html_path.read_text(encoding="utf-8", errors="ignore")
        for m in nav_re.finditer(html_blob):
            n = m.group(1).strip()
            loc = m.group(2).strip()
            # Strip " Rematch" suffix — use the base trainer's first appearance.
            base = re.sub(r"\s+rematch$", "", n, flags=re.I).strip().title()
            redux_locs[base].append(loc)
        print(f"Nuzlocke Redux HTML: {sum(len(v) for v in redux_locs.values())} "
              f"name->loc hints for {len(redux_locs)} unique trainers")

    # 5) Match parsed trainers to runtime IDs. Each parsed block claims its
    #    own runtime ID — duplicate names (multiple Bugsy fights, etc.) are
    #    assigned distinct IDs from the same name's candidate pool.
    parties: dict[str, dict] = {}
    unmatched: list[str] = []
    used_ids: set[int] = set()
    # Track block index per (name) so we know which match slot to claim.
    # Track how many times each name has been seen, so multiple fights for
    # the same trainer can each pick up their respective level-cap entry.
    name_seen: dict[str, int] = defaultdict(int)
    name_area_seen: dict[tuple[str, str], int] = defaultdict(int)
    for entry in raw_trainers:
        cls = entry["class"]
        name = entry["name"]
        party = entry["party"]
        party_size = len(party)
        # Resolve aliases so the spreadsheet's "Rival" finds Terry in
        # rr_trainers.json. The DISPLAY name keeps the spreadsheet value
        # so the dashboard shows "Rival" (which is what the player sees
        # in the in-game battle UI anyway).
        lookup_name = _NAME_ALIASES.get(name, name)
        rt_id = _pick_trainer_id(name_index, lookup_name, party_size, used_ids)
        if rt_id is None:
            unmatched.append(f"{cls} {name} ({entry['source']})")
            continue
        # If rt_id is already taken (same-name candidates exhausted), don't
        # clobber — the existing entry usually represents the same fight.
        if str(rt_id) in parties:
            continue
        used_ids.add(rt_id)
        # Pick this trainer's level cap. Two-tier lookup:
        #   1. If we know the area (from the sheet's 3-line header), try
        #      (name, area) first — same trainer can fight in the same area
        #      multiple times (rivals at Route 22 #1 / #2) and we want each
        #      visit's cap.
        #   2. Fall back to name-only, by encounter order.
        nm_t = name.title()
        entry_area = entry.get("area") or ""
        is_postgame = "Postgame" in (entry.get("source") or "")
        level_cap: int | None = None
        # Exact match takes priority — Trainer Order had a cap for this
        # (name, area) at this specific encounter index.
        if entry_area:
            area_caps = level_caps_by_area.get((nm_t, entry_area), [])
            area_idx = name_area_seen[(nm_t, entry_area)]
            if area_idx < len(area_caps):
                level_cap = area_caps[area_idx]
            name_area_seen[(nm_t, entry_area)] += 1
        if level_cap is None:
            cap_list = level_caps.get(nm_t, [])
            cap_idx = name_seen[nm_t]
            if cap_idx < len(cap_list):
                level_cap = cap_list[cap_idx]
        name_seen[nm_t] += 1
        # Fallbacks (in order of preference):
        #   1. Postgame entries → 100 (RR postgame is level-100 territory).
        #   2. Last cap seen for this name → for revisits the Trainer Order
        #      didn't enumerate (e.g. a Post-Surge Bugsy variant after the
        #      single Bugsy entry in Trainer Order).
        if level_cap is None:
            if is_postgame:
                level_cap = 100
            else:
                area_caps_fb = level_caps_by_area.get((nm_t, entry_area), [])
                if area_caps_fb:
                    level_cap = area_caps_fb[-1]
                else:
                    cap_list_fb = level_caps.get(nm_t, [])
                    if cap_list_fb:
                        level_cap = cap_list_fb[-1]
        # Compute a "static fallback level" per mon. Relative-level mons
        # (those with level_offset set — "Max Level - N" / "Highest Lv -N")
        # are resolved dynamically at render time against the player's
        # current highest party mon. Here we just fill in a reasonable
        # static estimate so renderers that DON'T have player state (the
        # supplementary calc JS, JSON consumers, etc.) still see plausible
        # numbers. Resolution order:
        #   1. mon["level"] already > 0 (sheet listed a concrete number).
        #   2. level_offset present + level_cap known → cap + offset.
        #   3. Postgame entries (no cap) → 100.
        #   4. last-resort default → 50 (mid-game).
        # The level_offset field is preserved on the mon either way so the
        # dashboard renderer can recompute dynamically when the player's
        # current state is richer than this static fallback.
        is_postgame = "Postgame" in (entry.get("source") or "")
        effective_cap = level_cap if level_cap is not None else (
            100 if is_postgame else None
        )
        for mon in party:
            lv = mon.get("level")
            if isinstance(lv, int) and lv > 0:
                continue
            off = mon.get("level_offset")
            if isinstance(off, int) and effective_cap is not None:
                mon["level"] = max(1, effective_cap + off)
            elif effective_cap is not None:
                mon["level"] = effective_cap
            else:
                mon["level"] = 50
        record = {
            "name":   name,
            "class":  cls,
            "source": entry["source"],
            "party":  party,
        }
        if entry.get("area"):
            record["area"] = entry["area"]
        if level_cap is not None:
            record["level_cap"] = level_cap
        # Carry the calc Prep-tab key + fight-variant tag + sprite URL
        # through to the JSON. calc_label drives the dashboard's
        # Open-in-Calc button (it's the SETDEX key); fight_label is the
        # human-readable "WHEN you fight him" tag; sprite_url comes from
        # the `=IMAGE(...)` cell above each header in the spreadsheet.
        if entry.get("calc_label"):
            record["calc_label"] = entry["calc_label"]
        if entry.get("fight_label"):
            record["fight_label"] = entry["fight_label"]
        if entry.get("sprite_url"):
            record["sprite_url"] = entry["sprite_url"]
        parties[str(rt_id)] = record

    # 6) Build area → [trainer_ids]. Three sources, processed in priority
    # order, each only contributing what the previous sources LEFT BLANK:
    #    (a) Each parsed entry's own header area — definitive, since the
    #        spreadsheet author stated the location verbatim in the
    #        2-line "LOCATION\nROLE" or 3-line "LOCATION\nCLASS\nNAME"
    #        header. One area per trainer ID.
    #    (b) Nuzlocke Redux nav list ("Brock at Pewter City Gym") — useful
    #        ONLY for trainers (a) didn't tag. Crucial: we attach a redux
    #        location to ONE unassigned rt_id per name, not every same-name
    #        rt_id. Otherwise "Archer" at Mt. Moon would also land at
    #        Cerulean Cave just because both are in the redux nav.
    #    (c) Trainer Order — same one-unassigned-rt_id-per-name rule.
    trainers_by_area: dict[str, list[int]] = defaultdict(list)
    assigned_area: dict[int, str] = {}    # rt_id → primary area

    def _attach(area: str, rt_id: int, *, primary: bool) -> None:
        if not area:
            return
        if rt_id in trainers_by_area[area]:
            return
        trainers_by_area[area].append(rt_id)
        if primary and rt_id not in assigned_area:
            assigned_area[rt_id] = area

    # (a) Header-derived areas. Each trainer attaches to exactly ONE area
    # (the one its own sheet header stated).
    for rt_id_str, info in parties.items():
        if info.get("area"):
            _attach(info["area"], int(rt_id_str), primary=True)

    # Build a queue of unassigned ids per (Title-cased name) so secondary
    # sources can hand out one location at a time.
    unassigned_by_name: dict[str, list[int]] = defaultdict(list)
    for rt_id_str, info in parties.items():
        if int(rt_id_str) in assigned_area:
            continue
        unassigned_by_name[info["name"].title()].append(int(rt_id_str))

    def _claim_unassigned(name_title: str) -> int | None:
        """Pop the next unassigned rt_id for this name, or None."""
        bucket = unassigned_by_name.get(name_title) or []
        if not bucket:
            return None
        return bucket.pop(0)

    # (b) Nuzlocke Redux — one rt_id per (name, redux_location) pair.
    for name, locs in redux_locs.items():
        for loc in locs:
            area_id = _normalize_area_name(loc)
            if not area_id:
                continue
            rt_id = _claim_unassigned(name)
            if rt_id is not None:
                _attach(area_id, rt_id, primary=True)

    # (c) Trainer Order — fills in anything redux didn't cover.
    for trainer_name, area_id, _is_optional in order_pairs:
        if not area_id:
            continue
        rt_id = _claim_unassigned(trainer_name.title())
        if rt_id is not None:
            _attach(area_id, rt_id, primary=True)

    # 6) Serialize.
    out_doc = {
        "_note": ("Auto-generated from the Pokémon Radical Red 4.1 community "
                  "trainer spreadsheet (Rudo2204 et al). Run "
                  "`python tools/gen_rr_priority_trainers.py` to regenerate. "
                  "The spreadsheet is canonical: sheet entries win over calc "
                  "data for both the displayed party and the calc_label used "
                  "by the dashboard's Open-in-Calc button. Trainer IDs are "
                  "1-based runtime IDs (matching the in-game trainer_id "
                  "reported by Lua); the corresponding 0-based index in "
                  "rr_trainers.json is `id - 1`. milestones is the Main "
                  "tab's level-cap progression; pre[name] is the cap when "
                  "approaching that fight, post[name] is the cap after "
                  "clearing it (= the next pre entry)."),
        "milestones": {
            "order": [{"name": n, "cap": c} for n, c in main_milestones],
            "pre":   main_pre_caps,
            "post":  main_post_caps,
        },
        "trainers_by_area": {k: sorted(v) for k, v in sorted(trainers_by_area.items())},
        "parties":          dict(sorted(parties.items(), key=lambda kv: int(kv[0]))),
    }

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        json.dump(out_doc, f, indent=2, ensure_ascii=False)

    # 7) Emit slink_priority.js — a supplementary calc SETDEX containing
    # every sheet-canonical fight, keyed by the synthesised calc_label so
    # the calc's Prep tab can find them by that exact name. Loaded by both
    # /calc/normal.html and /calc/hardcore.html immediately after the
    # vanilla normal.js / hardcore.js, so the new entries merge into
    # window.SETDEX_SV without overwriting existing keys.
    supp = _build_supplement_setdex(parties)
    supp_js = _render_supplement_js(supp)
    for path in (_CALC_SUPP_SRC_PATH, _CALC_SUPP_DIST_PATH):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(supp_js, encoding="utf-8")

    print()
    print(f"Wrote {out}")
    print(f"  trainers_by_area : {len(trainers_by_area)} areas")
    print(f"  parties          : {len(parties)} trainers")
    print(f"  unmatched        : {len(unmatched)}")
    print(f"Wrote {_CALC_SUPP_SRC_PATH} ({len(supp)} species, "
          f"{sum(len(v) for v in supp.values())} sets)")
    print(f"Wrote {_CALC_SUPP_DIST_PATH}")
    if unmatched[:10]:
        print(f"  first 10 unmatched:")
        for u in unmatched[:10]:
            print(f"    - {u}")
    return 0


def _build_supplement_setdex(parties: dict[str, dict]) -> dict[str, dict]:
    """Build {species: {calc_label: set_data}} from the matched parties.

    set_data has the same shape the calc consumes (nature, ability, level,
    item, moves, ivs, evs). Missing fields are filled with sensible RR
    defaults (max IVs, empty EVs) so the calc renders something usable
    even when the sheet only stated species/level/ability/item.
    """
    supp: dict[str, dict] = {}
    for info in parties.values():
        label = info.get("calc_label") or ""
        if not label:
            continue
        # Skip calc-fallback labels that include "Set N" — those are the
        # original (non-sheet) labels and they already exist in normal.js.
        # Only emit synthesised sheet labels.
        if re.search(r"\bSet\s+\d+\s*$", label):
            continue
        for mon in info.get("party") or []:
            species = mon.get("species") or ""
            if not species:
                continue
            level = mon.get("level")
            if not isinstance(level, int) or level <= 0:
                level = 50
            set_data: dict = {
                "nature":  mon.get("nature")  or "Hardy",
                "ability": mon.get("ability") or "",
                "level":   level,
                "moves":   [m for m in (mon.get("moves") or []) if m and m != "-"],
                "ivs":     {"hp":31,"at":31,"df":31,"sp":31,"sa":31,"sd":31},
            }
            item = (mon.get("item") or "").strip()
            if item and item.lower() not in ("none", "no item", "-"):
                set_data["item"] = item
            evs = mon.get("evs") or {}
            if evs:
                set_data["evs"] = dict(evs)
            supp.setdefault(species, {})[label] = set_data
    return supp


def _render_supplement_js(supp: dict[str, dict]) -> str:
    """Render the SETDEX supplement as an immediately-invoked JS function.

    The fragment expects window.SETDEX_SV to already be defined (i.e.
    normal.js / hardcore.js loaded first); it merges new entries IN PLACE
    without overwriting any existing keys, so we never trample a hand-tuned
    calc set even if our synthesised label happens to collide.
    """
    body = json.dumps(supp, indent=2, ensure_ascii=False)
    return (
        "// AUTO-GENERATED by tools/gen_rr_priority_trainers.py — do not edit.\n"
        "// Augments window.SETDEX_SV with priority-trainer sets transcribed\n"
        "// from the canonical RR boss spreadsheet. Loaded after normal.js /\n"
        "// hardcore.js so it can extend whichever setdex the page picked.\n"
        "(function () {\n"
        "  var ADD = " + body + ";\n"
        "  if (typeof window === 'undefined' || !window.SETDEX_SV) return;\n"
        "  Object.keys(ADD).forEach(function (sp) {\n"
        "    if (!window.SETDEX_SV[sp]) window.SETDEX_SV[sp] = {};\n"
        "    Object.keys(ADD[sp]).forEach(function (lbl) {\n"
        "      // Don't trample a hand-tuned set: only add when absent.\n"
        "      if (!window.SETDEX_SV[sp][lbl]) {\n"
        "        window.SETDEX_SV[sp][lbl] = ADD[sp][lbl];\n"
        "      }\n"
        "    });\n"
        "  });\n"
        "})();\n"
    )


if __name__ == "__main__":
    sys.exit(main())
